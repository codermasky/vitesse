from typing import List, Dict, Any, Union, Callable
import structlog
from pathlib import Path
import base64
import io
from PIL import Image
from langchain_core.messages import HumanMessage
from app.services.llm_provider import LLMProviderService
from app.core.langfuse_client import (
    trace_llm_call,
    record_llm_call_result,
    is_langfuse_enabled,
)

try:
    from langchain_text_splitters import RecursiveCharacterTextSplitter
except ImportError:
    from langchain.text_splitter import RecursiveCharacterTextSplitter
from langchain_community.vectorstores import FAISS

# import pinecone - REMOVED
# from pinecone import Pinecone - REMOVED

from app.core.config import settings

logger = structlog.get_logger(__name__)


class DocumentProcessor:
    """Service for processing and extracting text from documents."""

    def __init__(self):
        self.text_splitter = RecursiveCharacterTextSplitter(
            chunk_size=1000,
            chunk_overlap=200,
            length_function=len,
            separators=["\n\n", "\n", " ", ""],
        )

    async def describe_image(self, image_data: bytes, file_type: str = "jpeg") -> str:
        """Generate a description of an image using a vision model."""
        if not settings.ENABLE_VISION:
            return ""

        try:
            vision_model = await LLMProviderService.get_vision_model()

            # Convert bytes to base64
            b64_image = base64.b64encode(image_data).decode("utf-8")

            # Create message for vision model
            message = HumanMessage(
                content=[
                    {
                        "type": "text",
                        "text": "Describe this image in detail, focusing on any text, charts, data, or structural information relevant for a business context.",
                    },
                    {
                        "type": "image_url",
                        "image_url": {
                            "url": f"data:image/{file_type};base64,{b64_image}"
                        },
                    },
                ]
            )

            trace_ctx = None
            if is_langfuse_enabled():
                trace_ctx = trace_llm_call(
                    name="Vision Document Analysis",
                    model=getattr(vision_model, "model_name", "gpt-4o-vision"),
                    metadata={"content_type": f"image/{file_type}"},
                )

            if trace_ctx:
                async with trace_ctx as span:
                    response = await vision_model.ainvoke([message])
                    if span:
                        record_llm_call_result(
                            span, output=response.content if response else ""
                        )
            else:
                response = await vision_model.ainvoke([message])
            if not response or not response.content:
                logger.warning("Vision model returned empty response")
                return ""
            return f"[Image Description: {response.content}]\n"

        except AttributeError as ae:
            # Catch specific LangChain error when response is None
            if "'NoneType' object has no attribute 'model_dump'" in str(ae):
                logger.warning("Vision model returned None (caught AttributeError)")
                return ""
            logger.error("Image description failed with AttributeError", error=str(ae))
            return ""
        except Exception as e:
            logger.error("Image description failed", error=str(e))
            return "[Image Description: Failed to generate description]\n"

    async def extract_images_from_pdf(self, file_path: str) -> str:
        """Extract and describe images from PDF file."""
        image_descriptions = ""
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(file_path)
            for i, page in enumerate(reader.pages):
                if "/XObject" in page["/Resources"]:
                    xObject = page["/Resources"]["/XObject"].get_object()
                    for obj in xObject:
                        if xObject[obj]["/Subtype"] == "/Image":
                            try:
                                size = (xObject[obj]["/Width"], xObject[obj]["/Height"])
                                data = xObject[obj].get_data()
                                # Basic filter to skip tiny icons/lines
                                if size[0] < 100 or size[1] < 100:
                                    continue

                                description = await self.describe_image(data)
                                image_descriptions += (
                                    f"\n--- Image on Page {i + 1} ---\n{description}\n"
                                )
                            except Exception as img_e:
                                logger.warning(
                                    f"Failed to process image on page {i + 1}",
                                    error=str(img_e),
                                )
                                continue

            return image_descriptions
        except Exception as e:
            logger.error("PDF image extraction failed", error=str(e))
            return ""

    async def extract_text_with_metadata(self, file_path: str) -> Dict[str, Any]:
        """Extract text from PDF with metadata using PyMuPDF for better OCR and coordinate mapping."""
        if not file_path.lower().endswith(".pdf"):
            logger.warning(
                "Enhanced extraction only supported for PDF files", file_path=file_path
            )
            return {
                "text": "",
                "coordinate_mappings": [],
                "page_count": 0,
                "extraction_method": "skipped",
            }

        try:
            import fitz  # PyMuPDF

            with fitz.open(file_path) as doc:
                text_content = ""
                coordinate_mappings = []
                page_count = len(doc)

                for page_num in range(page_count):
                    page = doc.load_page(page_num)

                    # Enhanced text extraction with better OCR settings for financial docs
                    ocr_flags = (
                        fitz.TEXT_PRESERVE_LIGATURES
                        | fitz.TEXT_PRESERVE_WHITESPACE
                        | fitz.TEXT_INHIBIT_SPACES
                    )
                    text_instances = page.get_text("dict", flags=ocr_flags)

                    for block in text_instances.get("blocks", []):
                        if "lines" in block:
                            for line in block["lines"]:
                                for span in line["spans"]:
                                    text = span["text"].strip()
                                    if text:
                                        # Clean up common OCR errors in financial documents
                                        text = self._clean_financial_text(text)
                                        text_content += text + " "

                                        # Store coordinate mapping with enhanced metadata
                                        bbox = span["bbox"]  # [x0, y0, x1, y1]
                                        coordinate_mappings.append(
                                            {
                                                "text": text,
                                                "page": page_num + 1,
                                                "bbox": bbox,
                                                "font_size": span.get("size", 0),
                                                "font_name": span.get("font", ""),
                                                "color": span.get("color", 0),
                                                "flags": span.get("flags", 0),
                                            }
                                        )

            # Move image descriptions OUTSIDE fitz context to avoid handles conflict
            image_text = await self.extract_images_from_pdf(file_path)
            if image_text:
                text_content += "\n\n=== Extracted Image Contents ===\n" + image_text

            logger.info(
                f"Enhanced PDF extraction completed: {len(text_content)} chars, {len(coordinate_mappings)} text spans",
                file_path=file_path,
            )

            return {
                "text": text_content,
                "coordinate_mappings": coordinate_mappings,
                "page_count": page_count,
                "extraction_method": "pymupdf",
            }

        except ImportError:
            logger.warning("PyMuPDF not available, falling back to PyPDF2")
            text = await self.extract_text_from_pdf(file_path)
            return {
                "text": text,
                "coordinate_mappings": [],
                "page_count": 0,
                "extraction_method": "pypdf2",
            }
        except Exception as e:
            logger.error(
                "Enhanced PDF extraction failed", error=str(e), file_path=file_path
            )
            return {
                "text": "",
                "coordinate_mappings": [],
                "page_count": 0,
                "extraction_method": "failed",
            }

    def _clean_financial_text(self, text: str) -> str:
        """Clean common OCR errors in financial documents."""
        # Common OCR corrections for financial documents
        corrections = {
            "l": "1",  # lowercase L to 1
            "I": "1",  # uppercase I to 1
            "O": "0",  # uppercase O to 0
            "|": "1",  # pipe to 1
            "¢": "6",  # cent symbol to 6
            "§": "5",  # section symbol to 5
            "€": "6",  # euro symbol to 6
            "£": "6",  # pound symbol to 6
            "¥": "4",  # yen symbol to 4
        }

        # Apply corrections
        for wrong, correct in corrections.items():
            text = text.replace(wrong, correct)

        # Fix common financial abbreviations
        text = text.replace("c o g s", "cogs")
        text = text.replace("g a", "ga")
        text = text.replace("e b i t d a", "ebitda")
        text = text.replace("p l", "pl")
        text = text.replace("b s", "bs")

        return text

    async def extract_structured_data_from_pdf(
        self, file_path: str, doc_type: str = "unknown"
    ) -> Dict[str, Any]:
        """Extract structured financial data with coordinates for Click-to-Source functionality."""
        try:
            import fitz  # PyMuPDF

            with fitz.open(file_path) as doc:
                structured_data = {
                    "doc_type": doc_type,
                    "pages": [],
                    "financial_data": {},
                    "coordinate_mappings": {},
                }

                # Define comprehensive patterns for different document types and schedules
                patterns = {
                    "tax_return": {
                        "revenue": [
                            r"gross receipts",
                            r"total income",
                            r"gross revenue",
                            r"business income",
                            r"ordinary income",
                            r"rental income",
                        ],
                        "expenses": [
                            r"total expenses",
                            r"cost of goods sold",
                            r"cogs",
                            r"operating expenses",
                            r"business expenses",
                            r"deductions",
                        ],
                        "assets": [
                            r"total assets",
                            r"assets",
                            r"property.?plant.?equipment",
                            r"buildings",
                            r"land",
                            r"equipment",
                            r"machinery",
                        ],
                        "liabilities": [
                            r"total liabilities",
                            r"liabilities",
                            r"loans payable",
                            r"mortgages",
                            r"notes payable",
                            r"accounts payable",
                        ],
                        "equity": [
                            r"owner.?s equity",
                            r"retained earnings",
                            r"partner.?s capital",
                            r"stockholder.?s equity",
                            r"member.?s capital",
                        ],
                        "tax_year": [r"tax year", r"year ended", r"fiscal year"],
                        # Schedule L (Balance Sheet) specific patterns
                        "schedule_l_assets": [
                            r"schedule l",
                            r"balance sheet",
                            r"assets.*beginning",
                            r"assets.*end",
                            r"cash.*beginning",
                            r"cash.*end",
                        ],
                        "schedule_l_liabilities": [
                            r"liabilities.*beginning",
                            r"liabilities.*end",
                            r"loans.*beginning",
                            r"loans.*end",
                        ],
                        # Schedule M-1 (Reconciliation) specific patterns
                        "schedule_m1": [
                            r"schedule m-1",
                            r"reconciliation",
                            r"book.*tax",
                            r"net income.*book",
                            r"net income.*tax",
                        ],
                        # Schedule M-3 (for larger businesses) specific patterns
                        "schedule_m3": [
                            r"schedule m-3",
                            r"reconciliation.*financial",
                            r"financial.*reconciliation",
                        ],
                    },
                    "profit_loss": {
                        "revenue": [
                            r"total revenue",
                            r"gross sales",
                            r"net sales",
                            r"service revenue",
                            r"product revenue",
                            r"other income",
                        ],
                        "cogs": [
                            r"cost of goods sold",
                            r"cogs",
                            r"cost of sales",
                            r"direct costs",
                            r"product costs",
                        ],
                        "operating_expenses": [
                            r"operating expenses",
                            r"total expenses",
                            r"general.*admin",
                            r"g.?a.?expenses",
                            r"selling.*expenses",
                            r"admin.*expenses",
                        ],
                        "net_income": [
                            r"net income",
                            r"net profit",
                            r"profit before tax",
                            r"operating income",
                            r"ebit",
                            r"earnings before interest",
                        ],
                        "depreciation": [
                            r"depreciation",
                            r"amortization",
                            r"depreciation.*expense",
                        ],
                        "interest_expense": [
                            r"interest expense",
                            r"interest paid",
                            r"finance costs",
                        ],
                    },
                    "balance_sheet": {
                        "assets": [
                            r"total assets",
                            r"assets",
                            r"current assets",
                            r"fixed assets",
                            r"property.?plant.?equipment",
                            r"intangible assets",
                        ],
                        "liabilities": [
                            r"total liabilities",
                            r"liabilities",
                            r"current liabilities",
                            r"long.?term.*liabilities",
                            r"debt",
                            r"loans",
                        ],
                        "equity": [
                            r"total equity",
                            r"owner.?s equity",
                            r"stockholder.?s equity",
                            r"retained earnings",
                            r"capital",
                        ],
                        "cash": [r"cash.*equivalents", r"cash.*bank", r"cash.*balance"],
                    },
                    "rent_roll": {
                        "tenant_info": [
                            r"tenant",
                            r"lease",
                            r"rent",
                            r"occupant",
                            r"unit",
                        ],
                        "rental_income": [
                            r"rental income",
                            r"rent collected",
                            r"lease income",
                            r"monthly rent",
                            r"annual rent",
                        ],
                        "vacancy": [r"vacant", r"vacancy", r"empty", r"available"],
                    },
                }

                doc_patterns = patterns.get(doc_type, {})

                def _convert_bbox(rect, page_height):
                    """Convert PyMuPDF [x0, y0, x1, y1] to [x, y_from_bottom, w, h] for frontend."""
                    x0, y0, x1, y1 = rect
                    return [
                        x0,
                        max(0, page_height - y1),
                        max(0, x1 - x0),
                        max(0, y1 - y0),
                    ]

                for page_num in range(len(doc)):
                    page = doc.load_page(page_num)
                    page_height = page.rect.height
                    page_data = {
                        "page_number": page_num + 1,
                        "text_blocks": [],
                        "tables": [],
                    }

                    # Extract text blocks with coordinates
                    text_instances = page.get_text("dict")
                    for block in text_instances.get("blocks", []):
                        if "lines" in block:
                            bbox = _convert_bbox(block["bbox"], page_height)
                            block_text = ""

                            # Refined extraction: check each line for pattern matches
                            for line in block["lines"]:
                                line_text = ""
                                line_bbox = _convert_bbox(line["bbox"], page_height)
                                for span in line["spans"]:
                                    line_text += span["text"]
                                    block_text += span["text"]

                                if line_text.strip():
                                    # Check for financial data patterns at the LINE level for precision
                                    for field_type, regexes in doc_patterns.items():
                                        import re

                                        for regex in regexes:
                                            if re.search(regex, line_text.lower()):
                                                if (
                                                    field_type
                                                    not in structured_data[
                                                        "financial_data"
                                                    ]
                                                ):
                                                    structured_data["financial_data"][
                                                        field_type
                                                    ] = []

                                                # Use line_bbox for maximum precision
                                                structured_data["financial_data"][
                                                    field_type
                                                ].append(
                                                    {
                                                        "text": line_text.strip(),
                                                        "bbox": line_bbox,
                                                        "page": page_num + 1,
                                                    }
                                                )

                            if block_text.strip():
                                page_data["text_blocks"].append(
                                    {"text": block_text.strip(), "bbox": bbox}
                                )

                    # Extract tables
                    tabs = page.find_tables()
                    for tab in tabs:
                        table_data = tab.extract()
                        if table_data:
                            page_data["tables"].append(
                                {
                                    "data": table_data,
                                    "bbox": _convert_bbox(tab.bbox, page_height),
                                }
                            )

                    structured_data["pages"].append(page_data)

            # Create coordinate mappings for Click-to-Source
            for field_type, entries in structured_data["financial_data"].items():
                structured_data["coordinate_mappings"][field_type] = entries

            logger.info(
                f"Structured data extraction completed for {doc_type} with line-level precision",
                file_path=file_path,
                pages=len(structured_data["pages"]),
            )

            return structured_data

        except ImportError:
            logger.warning("PyMuPDF not available for structured extraction")
            return {
                "doc_type": doc_type,
                "pages": [],
                "financial_data": {},
                "coordinate_mappings": {},
            }
        except Exception as e:
            logger.error(
                "Structured data extraction failed", error=str(e), file_path=file_path
            )
            return {
                "doc_type": doc_type,
                "pages": [],
                "financial_data": {},
                "coordinate_mappings": {},
            }

    async def extract_text_from_pdf(self, file_path: str) -> str:
        """Extract text from PDF file."""
        try:
            from PyPDF2 import PdfReader

            reader = PdfReader(file_path)
            text = ""
            for page in reader.pages:
                text += page.extract_text() + "\n"

            # Add image descriptions
            logger.info("Extracting images from PDF...", file_path=file_path)
            image_text = await self.extract_images_from_pdf(file_path)
            if image_text:
                text += "\n\n=== Extracted Image Contents ===\n" + image_text

            logger.info(
                "PDF text extraction completed",
                file_path=file_path,
                pages=len(reader.pages),
            )
            return text

        except Exception as e:
            logger.error("PDF extraction failed", error=str(e), file_path=file_path)
            return ""

    async def extract_text_from_docx(self, file_path: str) -> str:
        """Extract text from DOCX file."""
        try:
            from docx import Document

            doc = Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"

            logger.info("DOCX text extraction completed", file_path=file_path)
            return text

        except Exception as e:
            logger.error("DOCX extraction failed", error=str(e), file_path=file_path)
            return ""

    async def extract_text_from_excel(self, file_path: str) -> str:
        """Extract text from Excel file."""
        try:
            if file_path.lower().endswith(".xlsx"):
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
                text = ""
                for sheet in wb.worksheets:
                    for row in sheet.iter_rows(values_only=True):
                        row_text = [
                            str(cell).strip()
                            for cell in row
                            if cell is not None and str(cell).strip()
                        ]
                        if row_text:
                            # Use pipe separator to preserve column structure for RAG
                            # e.g. "Question | Answer | Category"
                            text += " | ".join(row_text) + "\n"
            elif file_path.lower().endswith(".xls"):
                import xlrd

                wb = xlrd.open_workbook(file_path)
                text = ""
                for sheet in wb.sheets():
                    for row_idx in range(sheet.nrows):
                        row = sheet.row(row_idx)
                        row_text = [
                            str(cell.value).strip()
                            for cell in row
                            if cell.value is not None and str(cell.value).strip()
                        ]
                        if row_text:
                            text += " | ".join(row_text) + "\n"
            else:
                raise ValueError("Unsupported Excel file format")

            logger.info("Excel text extraction completed", file_path=file_path)
            return text

        except Exception as e:
            logger.error("Excel extraction failed", error=str(e), file_path=file_path)
            return ""

    async def extract_financial_data_from_excel(
        self, file_path: str, doc_type: str = "spreadsheet"
    ) -> Dict[str, Any]:
        """Extract structured financial data with cell coordinates for Click-to-Source functionality."""
        try:
            structured_data = {
                "doc_type": doc_type,
                "sheets": [],
                "financial_data": {},
                "coordinate_mappings": {},
            }

            if file_path.lower().endswith(".xlsx"):
                import openpyxl

                wb = openpyxl.load_workbook(file_path, data_only=True)
                sheets = wb.worksheets
            elif file_path.lower().endswith(".xls"):
                import xlrd

                wb = xlrd.open_workbook(file_path)
                sheets = wb.sheets()
            else:
                raise ValueError("Unsupported Excel file format")

            # Define patterns for financial data
            patterns = {
                "revenue": [
                    r"gross revenue",
                    r"gross receipts",
                    r"total income",
                    r"business income",
                    r"rental income",
                    r"ordinary income",
                    r"sales",
                ],
                "expenses": [
                    r"total expenses",
                    r"cost of goods sold",
                    r"cogs",
                    r"operating expenses",
                    r"business expenses",
                    r"deductions",
                    r"salaries",
                    r"rent",
                    r"utilities",
                ],
                "assets": [
                    r"total assets",
                    r"assets",
                    r"property",
                    r"equipment",
                    r"machinery",
                    r"buildings",
                    r"land",
                    r"cash",
                    r"accounts receivable",
                ],
                "liabilities": [
                    r"total liabilities",
                    r"liabilities",
                    r"loans payable",
                    r"mortgages",
                    r"notes payable",
                    r"accounts payable",
                    r"debt",
                ],
                "equity": [
                    r"owner.?s equity",
                    r"retained earnings",
                    r"partner.?s capital",
                    r"stockholder.?s equity",
                    r"member.?s capital",
                ],
                "tax_year": [r"tax year", r"year ended", r"fiscal year"],
            }

            doc_patterns = patterns.get(doc_type, patterns)

            for sheet_idx, sheet in enumerate(sheets):
                sheet_name = (
                    sheet.title if hasattr(sheet, "title") else f"Sheet{sheet_idx + 1}"
                )
                sheet_data = {
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_idx,
                    "cells": [],
                }

                if file_path.lower().endswith(".xlsx"):
                    # openpyxl
                    for row_idx, row in enumerate(
                        sheet.iter_rows(values_only=True), start=1
                    ):
                        for col_idx, cell_value in enumerate(row, start=1):
                            if cell_value is not None and str(cell_value).strip():
                                cell_text = str(cell_value).strip()
                                sheet_data["cells"].append(
                                    {
                                        "row": row_idx,
                                        "col": col_idx,
                                        "text": cell_text,
                                    }
                                )

                                # Check for financial data patterns
                                for field_type, regexes in doc_patterns.items():
                                    import re

                                    for regex in regexes:
                                        if re.search(regex, cell_text.lower()):
                                            if (
                                                field_type
                                                not in structured_data["financial_data"]
                                            ):
                                                structured_data["financial_data"][
                                                    field_type
                                                ] = []
                                            structured_data["financial_data"][
                                                field_type
                                            ].append(
                                                {
                                                    "text": cell_text,
                                                    "coordinates": [
                                                        row_idx,
                                                        col_idx,
                                                    ],  # [row, col] for Excel
                                                    "sheet": sheet_idx,
                                                    "field": field_type,
                                                }
                                            )
                else:
                    # xlrd
                    for row_idx in range(sheet.nrows):
                        for col_idx in range(sheet.ncols):
                            cell_value = sheet.cell_value(row_idx, col_idx)
                            if cell_value is not None and str(cell_value).strip():
                                cell_text = str(cell_value).strip()
                                sheet_data["cells"].append(
                                    {
                                        "row": row_idx + 1,  # 1-based
                                        "col": col_idx + 1,
                                        "text": cell_text,
                                    }
                                )

                                # Check for financial data patterns
                                for field_type, regexes in doc_patterns.items():
                                    import re

                                    for regex in regexes:
                                        if re.search(regex, cell_text.lower()):
                                            if (
                                                field_type
                                                not in structured_data["financial_data"]
                                            ):
                                                structured_data["financial_data"][
                                                    field_type
                                                ] = []
                                            structured_data["financial_data"][
                                                field_type
                                            ].append(
                                                {
                                                    "text": cell_text,
                                                    "coordinates": [
                                                        row_idx + 1,
                                                        col_idx + 1,
                                                    ],  # [row, col] for Excel
                                                    "sheet": sheet_idx,
                                                    "field": field_type,
                                                }
                                            )

                structured_data["sheets"].append(sheet_data)

            # Create coordinate mappings for Click-to-Source
            for field_type, entries in structured_data["financial_data"].items():
                structured_data["coordinate_mappings"][field_type] = entries

            logger.info(
                f"Excel structured data extraction completed",
                file_path=file_path,
                sheets=len(structured_data["sheets"]),
            )

            return structured_data

        except Exception as e:
            logger.error(
                "Excel structured data extraction failed",
                error=str(e),
                file_path=file_path,
            )
            return {
                "doc_type": doc_type,
                "sheets": [],
                "financial_data": {},
                "coordinate_mappings": {},
            }

    async def process_document(self, file_path: str, file_type: str) -> str:
        """Process document and extract text based on file type."""
        file_ext = file_type.lower()

        if file_ext == "pdf":
            return await self.extract_text_from_pdf(file_path)
        elif file_ext in ["docx", "doc"]:
            return await self.extract_text_from_docx(file_path)
        elif file_ext in ["xlsx", "xls"]:
            return await self.extract_text_from_excel(file_path)
        elif file_ext in ["jpg", "jpeg", "png", "bmp", "tiff"]:
            try:
                with open(file_path, "rb") as split_file:
                    header = split_file.read(4)
                    split_file.seek(0)
                    image_data = split_file.read()

                return await self.describe_image(
                    image_data, file_type=file_ext if file_ext != "jpg" else "jpeg"
                )
            except Exception as e:
                logger.error(
                    "Image processing failed", error=str(e), file_path=file_path
                )
                return ""
        else:
            # For other file types, try to read as plain text
            try:
                with open(file_path, "r", encoding="utf-8") as f:
                    return f.read()
            except Exception as e:
                logger.error(
                    "Text extraction failed", error=str(e), file_path=file_path
                )
                return ""


class VectorizationService:
    """Service for vectorizing documents and managing vector database."""

    def __init__(self):
        self.document_processor = DocumentProcessor()
        self.embeddings = self._get_embeddings()
        self.vector_store = self._get_vector_store()

    def _get_embeddings(self):
        """Get embeddings model."""
        try:
            from langchain_huggingface import HuggingFaceEmbeddings

            return HuggingFaceEmbeddings(
                model_name="all-MiniLM-L6-v2", cache_folder="/models"
            )
        except ImportError:
            logger.warning(
                "HuggingFace embeddings not available, falling back to OpenAI"
            )
            if settings.OPENAI_API_KEY:
                from langchain_openai import OpenAIEmbeddings

                return OpenAIEmbeddings(
                    model="text-embedding-ada-002",
                    openai_api_key=settings.OPENAI_API_KEY,
                )
            raise ValueError("No embedding model available")

    def _get_vector_store(self):
        """Get vector store based on configuration."""
        if settings.VECTOR_DB_TYPE == "pinecone":
            return self._get_pinecone_store()
        elif settings.VECTOR_DB_TYPE == "faiss":
            return self._get_faiss_store()
        else:
            raise ValueError(f"Unsupported vector DB type: {settings.VECTOR_DB_TYPE}")

    def _get_pinecone_store(self):
        """Initialize Pinecone vector store."""
        from langchain_pinecone import PineconeVectorStore
        from pinecone import Pinecone

        if not settings.PINECONE_API_KEY:
            raise ValueError("Pinecone API key not configured")

        pc = Pinecone(api_key=settings.PINECONE_API_KEY)

        # Check if index exists, create if not
        if settings.PINECONE_INDEX_NAME not in pc.list_indexes().names():
            pc.create_index(
                name=settings.PINECONE_INDEX_NAME,
                dimension=1536,  # OpenAI Ada-002 dimension
                metric="cosine",
            )

        return PineconeVectorStore(
            pinecone_api_key=settings.PINECONE_API_KEY,
            index_name=settings.PINECONE_INDEX_NAME,
            embedding=self.embeddings,
        )

    def _get_faiss_store(self):
        """Initialize FAISS vector store."""
        # For FAISS, we'll create/load from local storage
        index_path = Path(settings.UPLOAD_DIR) / "faiss_index"
        index_path.mkdir(parents=True, exist_ok=True)

        try:
            return FAISS.load_local(
                str(index_path), self.embeddings, allow_dangerous_deserialization=True
            )
        except Exception:
            # Create new empty index
            return FAISS.from_texts([""], self.embeddings)

    async def process_and_vectorize_document(
        self,
        file_path: str,
        file_type: str,
        document_id: str,
        metadata: Dict[str, Any] = None,
        db_session=None,
    ) -> Dict[str, Any]:
        """Process document, extract text, and vectorize it with status tracking."""
        from app.services.metadata_service import get_metadata_service
        from app.models.document import ExtractionStatus

        metadata_service = None
        if db_session:
            metadata_service = get_metadata_service(db_session)
            # Update status to PROCESSING
            await metadata_service.update_extraction_status(
                document_id, ExtractionStatus.PROCESSING
            )

        try:
            # Extract text
            text = await self.document_processor.process_document(file_path, file_type)

            if not text.strip():
                if metadata_service:
                    await metadata_service.update_extraction_status(
                        document_id,
                        ExtractionStatus.FAILED,
                        error="No text extracted from document",
                    )
                return {
                    "success": False,
                    "error": "No text extracted from document",
                    "document_id": document_id,
                }

            # Split into chunks
            chunks = self.document_processor.text_splitter.split_text(text)

            # Prepare metadata and IDs for each chunk
            base_metadata = {
                "document_id": document_id,
                "file_type": file_type,
                "source": file_path,
                **(metadata or {}),
            }

            metadatas = []
            ids = []
            for i, chunk in enumerate(chunks):
                chunk_metadata = base_metadata.copy()
                chunk_metadata.update(
                    {
                        "chunk_id": f"{document_id}_chunk_{i}",
                        "chunk_index": i,
                        "total_chunks": len(chunks),
                    }
                )
                metadatas.append(chunk_metadata)
                ids.append(f"{document_id}_chunk_{i}")

            # Add to vector store
            # FAISS
            # Use add_texts directly instead of merge_from for better reliability
            self.vector_store.add_texts(chunks, metadatas=metadatas, ids=ids)

            # Save FAISS index
            index_path = Path(settings.UPLOAD_DIR) / "faiss_index"
            self.vector_store.save_local(str(index_path))

            # Update extraction status to COMPLETED
            if metadata_service:
                await metadata_service.update_extraction_status(
                    document_id,
                    ExtractionStatus.COMPLETED,
                    chunk_count=len(chunks),
                    text_length=len(text),
                    embedding_model="all-MiniLM-L6-v2",
                )

            logger.info(
                "Document vectorized successfully",
                document_id=document_id,
                chunks=len(chunks),
                text_length=len(text),
            )

            return {
                "success": True,
                "document_id": document_id,
                "chunks": len(chunks),
                "text_length": len(text),
                "metadata": base_metadata,
            }

        except Exception as e:
            # Update extraction status to FAILED
            if metadata_service:
                await metadata_service.update_extraction_status(
                    document_id, ExtractionStatus.FAILED, error=str(e)
                )

            logger.error(
                "Document vectorization failed",
                error=str(e),
                document_id=document_id,
                file_path=file_path,
            )
            return {"success": False, "error": str(e), "document_id": document_id}

    async def search_similar(
        self,
        query: str,
        k: int = 5,
        filter_metadata: Union[Dict[str, Any], Callable[[Dict[str, Any]], bool]] = None,
    ) -> List[Dict[str, Any]]:
        """Search for similar documents in the vector store."""
        try:
            # FAISS with metadata filtering
            # Note: FAISS in LangChain uses the 'filter' argument for metadata.
            # We must ensure filter_metadata contains at least one common item if provided.
            docs = self.vector_store.similarity_search(
                query, k=k, filter=filter_metadata
            )

            results = []
            for doc in docs:
                results.append(
                    {
                        "content": doc.page_content,
                        "metadata": doc.metadata,
                        "score": getattr(doc, "score", None),
                    }
                )

            logger.info(
                "Vector search completed",
                query_length=len(query),
                results_count=len(results),
            )
            return results

        except Exception as e:
            logger.error("Vector search failed", error=str(e), query=query[:100])
            return []

    async def delete_document(self, document_id: str, chunk_count: int = None) -> bool:
        """Delete document vectors from the store."""
        try:
            # FAISS
            # Attempt to delete by IDs if chunk_count is known or we can guess
            if not chunk_count:
                # If we don't know the count, we can't easily generate all IDs to delete in FAISS
                # without iterating. For now, we log a warning.
                # TODO: Implement full scan deletion or mapping lookups
                logger.warning(
                    "Deleting document without chunk_count is generic (may leave ghosts)",
                    document_id=document_id,
                )
                return True

            ids_to_delete = [f"{document_id}_chunk_{i}" for i in range(chunk_count)]

            if hasattr(self.vector_store, "delete"):
                result = self.vector_store.delete(ids_to_delete)
                logger.info("Deleted vectors", document_id=document_id, result=result)
                # Save after deletion
                index_path = Path(settings.UPLOAD_DIR) / "faiss_index"
                self.vector_store.save_local(str(index_path))
                return True
            else:
                logger.warning("Vector store does not support deletion")
                return False

        except Exception as e:
            logger.error(
                "Document deletion failed", error=str(e), document_id=document_id
            )
            return False

    async def revectorize_document(
        self,
        document_id: str,
        file_path: str,
        file_type: str,
        metadata: Dict[str, Any],
        chunk_count_for_deletion: int = None,
        db_session=None,
    ) -> Dict[str, Any]:
        """Re-vectorize a document (delete old -> add new) to update content/metadata."""
        logger.info("Re-vectorizing document", document_id=document_id)

        # 1. Delete old vectors (best effort)
        if chunk_count_for_deletion:
            await self.delete_document(
                document_id, chunk_count=chunk_count_for_deletion
            )

        # 2. Add as new
        return await self.process_and_vectorize_document(
            file_path=file_path,
            file_type=file_type,
            document_id=document_id,
            metadata=metadata,
            db_session=db_session,
        )

    async def get_document_stats(self) -> Dict[str, Any]:
        """Get statistics about the vector store."""
        try:
            # FAISS stats
            return {
                "index_type": "FAISS",
                "vectors": getattr(self.vector_store, "index", None)
                and self.vector_store.index.ntotal
                or 0,
            }

        except Exception as e:
            logger.error("Failed to get vector store stats", error=str(e))
            return {"error": str(e)}


_vectorization_service = None


def get_vectorization_service():
    """Get or create the global vectorization service instance."""
    global _vectorization_service
    if _vectorization_service is None:
        _vectorization_service = VectorizationService()
    return _vectorization_service
